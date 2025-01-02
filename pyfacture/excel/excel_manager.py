import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def update_excel(file_path, ticket_data):
    """
    Updates Excel file
    
    param: 
        file_path: path to excel file
        ticket_data: ticket data
    """
    if not os.path.exists(file_path):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Dépenses"

        sheet.append(["Date", "Product", "Price", "Total (€)"])
    else:
        workbook = load_workbook(file_path)
        sheet = workbook.active

    date = ticket_data["date"]
    total = ticket_data["total"]

    for product in ticket_data["products"]:
        sheet.append([date, product["name"], product["price"], ""])

    sheet.append([date, "", "", total])
    workbook.save(file_path)
